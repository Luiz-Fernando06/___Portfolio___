# Verificar se os valores estão pagos ou não e caso estejam, verificar a forma de pagamento

# com base em uma planilha, pegar o cpf e consultar no site disponibilizado para veificar se a conta está(ou não) paga

# Caso esteja paga, preencher na planilha de fachamento como "ok", caso contrario, informar que cfontinua pedente

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep

# 1 - entrar na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_cliente = planilha_clientes['Sheet1']

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
url = 'https://consultcpf-devaprender.netlify.app/'
driver.get(url)

planilha_fechamento = openpyxl.load_workbook('planilha fechamento(1).xlsx')
pagina_fechamento = planilha_fechamento['Sheet1']

# 7 - Repetir até chegar na ultima linha
for linha in pagina_cliente.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    # 2 - Entro no site https://consultcpf-devaprender.netlify.app/
    # colo o cpf do cliente e verifico se o pagamento está em dia
    sleep(1)
    campo_cpf = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(1)
    campo_cpf.clear()  # apaga oq esta no campo
    sleep(1)
    campo_cpf.send_keys(cpf)  # permite que eu escrava algo no campo
    sleep(1)

    # 3 - verificar se está "em dia" ou "atrasado"
    botao_consulta = driver.find_element(
        By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(4)
    botao_consulta.click()

    # 4 - se estiver em dia, pegar a data do pagamento e o metodo de pagamento
    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")

    if status.text == "em dia":
        # 6 - inserir essas novas informações(nome, valor, cpf, vencimento, status e caso esteja em dia, data pagamento, metodo pagamento(cartão ou boleto)) em uma nova planilha
        data_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentDate']")
        data_pagamento_limpo = data_pagamento.text.split(' ')[3]

        metodo_pagamento = driver.find_element(
            By.XPATH, "//p[@id='paymentMethod']")
        metodo_pagamento_limpo = metodo_pagamento.text.split(' ')[3]

        pagina_fechamento.append(
            [nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

        planilha_fechamento.save('planilha fechamento(1).xlsx')

    else:
        # 5 - Caso contrario(se estiver atrasado), colocar o status como pendente em uma nova planilha
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])

        planilha_fechamento.save('planilha fechamento(1).xlsx')

driver.quit()
